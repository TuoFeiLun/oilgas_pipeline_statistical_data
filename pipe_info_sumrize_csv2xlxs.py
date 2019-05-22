import csv
from pprint import pprint
from openpyxl import Workbook
from openpyxl import load_workbook

'''
直接读取csv文件，生成excel，不用改表头表尾，但文件必须包含材料，壁厚，外径  最好是只包含这几项

测试ok
测试文件


统计同一管径下，同一材料，不同的壁厚，
主要是一些公司人员在录入信息时输入错误，正常情况下 同一外径，材料，下壁厚最多2种，然而，现实情况并非如此，可能是最多时时10多种

'''

csvfilepath = r'E:\li\all-pipeline-info.csv'   # 读取原始数据的位置
save_path = r'E:\li\all-pipeline-info-1.xlsx'   # 存储位置
with open(csvfilepath, newline='') as csvfile:
    spamreader = csv.reader(csvfile, delimiter=' ', quotechar='|')
    all_list = list()
    for row in spamreader:
        all_list.append(row)

# pprint(all_list)

#def column_attribute():
material_letter = ''
outerdiameter_letter = ''
thickness_letter = ''
temp1 = list()
for i in range(len(all_list[2])):
    temp1.append(all_list[2][i])


titles = ','.join(temp1)
titles_list = titles.split(',')
remove_num = len(all_list[2])-len(all_list[3])
print(titles_list)
for i in range(remove_num):
    titles_list.remove('(mm)')

for i in range(1,len(titles_list)):

    title = titles_list[i]
    if title is None:
        continue
    if '材料' in title:
        material_letter = i
    elif '外径' in title:
        outerdiameter_letter = i
    elif '壁厚' in title:
        thickness_letter = i

print(material_letter,outerdiameter_letter,thickness_letter)
max_row = len(all_list)-6
start_row =4
alldict = {}
pprint(all_list)
print(all_list[max_row],all_list[max_row-1])
for i in range(start_row, max_row):


    temp2 = all_list[i][0].split(',')

    diameter = float(temp2[outerdiameter_letter])
    alldict[diameter] ={}

for i in range(start_row, max_row):

    temp2 = all_list[i][0].split(',')

    diameter = float(temp2[outerdiameter_letter])

    material = temp2[material_letter]
    if material =='':
        material = '无材料'
    alldict[diameter][material] = set()

for i in range(start_row, max_row):
    temp2 = all_list[i][0].split(',')

    diameter = float(temp2[outerdiameter_letter])

    material = temp2[material_letter]
    if material == '':
        material = '无材料'

    thickness = temp2[thickness_letter]

    alldict[diameter][material].add(thickness)

def sortedDictValues2(adict):   # 对dict排序 后，返回dict
    adict = sorted(adict.items(), key=lambda item: item[0])
    new_adict = dict()
    for i in range(len(adict)):
        new_adict[adict[i][0]] = adict[i][1]
    return new_adict
alldict = sortedDictValues2(alldict)
pprint(alldict)
alllist = list()
for key,value in alldict.items():

    for material, thickness in value.items():
        keylist = list()
        keylist.append(key)
        keylist.append(material)
        for only_thickness in thickness:
            keylist.append(only_thickness)

        alllist.append(keylist)



wb = Workbook(write_only=True)
ws_2 = wb.create_sheet()
biaotou = ['外径','材料信息','壁厚']
ws_2.append(biaotou)
for i in alllist:

    ws_2.append(i)

wb.save(save_path)
print('ok')
# ff = ['管道316373,,26.24,3.58,']
# fff = ff[0].split(',')
# if fff[1] == '':
#     print(repr(fff[1]))
