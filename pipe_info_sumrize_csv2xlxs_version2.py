import csv
from pprint import pprint
from openpyxl import Workbook
from openpyxl import load_workbook
import math

'''
这是统计了每种管道的长度 
直接读取csv文件，生成excel，不用改表头表尾，但文件必须包含材料，壁厚，外径  最好是只包含这几项
 
测试ok
测试文件

 r 'E:\19-2-25\test.csv' 
统计同一管径下，不同的壁厚 

2019 5-17 这里我们想添加获取一种管径的长度
名称	材料	外径 (mm)	壁厚 (mm)	粗糙度 (mm)	长度 (m)

测试 ok 
'''

csvfilepath = r'E:\\all-pipeline-info-2.csv'
save_path = r'E:\\sum-all-pipeline-info-2.xlsx'
with open(csvfilepath, newline='') as csvfile:
    spamreader = csv.reader(csvfile, delimiter=' ', quotechar='|')
    all_list = list()
    for row in spamreader:
        all_list.append(row)

# pprint(all_list)

#def column_attribute():
material_letter = ''
outerdiameter_letter = ''
thickness_letter = ''
length_letter = ''
temp1 = list()
for i in range(len(all_list[2])):
    temp1.append(all_list[2][i])

print(temp1)
titles = ','.join(temp1)
titles_list = titles.split(',')
# remove_num = len(all_list[2])-len(all_list[3])
print(titles_list)
# for i in range(remove_num):
#     titles_list.remove('(mm)')
i=0
while i < len(titles_list):
    if 'm' in titles_list[i]:
        titles_list.remove(titles_list[i])
    i += 1
print(titles_list)
for i in range(1,len(titles_list)):

    title = titles_list[i]
    if title is None:
        continue
    if '材料' in title:
        material_letter = i
    elif '外径' in title:
        outerdiameter_letter = i
    elif '壁厚' in title:
        thickness_letter = i
    elif '长度' in title:
        length_letter = i
print(material_letter,outerdiameter_letter,thickness_letter,length_letter)
max_row = len(all_list)-6
start_row =4
alldict = {}
pprint(all_list)
print(all_list[max_row],all_list[max_row-1])
for i in range(start_row, max_row):


    temp2 = all_list[i][0].split(',')

    diameter = float(temp2[outerdiameter_letter])
    alldict[diameter] ={}

for i in range(start_row, max_row):

    temp2 = all_list[i][0].split(',')

    diameter = float(temp2[outerdiameter_letter])

    material = temp2[material_letter]
    if material =='':
        material = '无材料'
    alldict[diameter][material] = {}   # 该字典 存储 壁厚：长度

for i in range(start_row, max_row):
    temp2 = all_list[i][0].split(',')

    diameter = float(temp2[outerdiameter_letter])

    material = temp2[material_letter]
    if material == '':
        material = '无材料'

    thickness = temp2[thickness_letter]
    pipe_length = float(temp2[length_letter])

    if thickness not in alldict[diameter][material]:

        alldict[diameter][material][thickness] = pipe_length
    else:
        alldict[diameter][material][thickness] += pipe_length

def sortedDictValues2(adict):   # 对dict排序 后，返回dict
    adict = sorted(adict.items(), key=lambda item: item[0])
    new_adict = dict()
    for i in range(len(adict)):
        new_adict[adict[i][0]] = adict[i][1]
    return new_adict
alldict = sortedDictValues2(alldict)
pprint(alldict)
alllist = list()
for key,value in alldict.items():

    for material, thickness in value.items():
        keylist = list()
        keylist.append(key)
        keylist.append(material)
        for only_thickness in thickness:
            keylist.append(str(only_thickness)+'-'+str(math.ceil(thickness[only_thickness])))

        alllist.append(keylist)



wb = Workbook(write_only=True)
ws_2 = wb.create_sheet()
biaotou = ['外径','材料信息','壁厚+长度']
ws_2.append(biaotou)
for i in alllist:

    ws_2.append(i)

wb.save(save_path)
print('ok')
# ff = ['管道316373,,26.24,3.58,']
# fff = ff[0].split(',')
# if fff[1] == '':
#     print(repr(fff[1]))
